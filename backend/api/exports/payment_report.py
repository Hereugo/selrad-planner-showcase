import logging
from datetime import datetime
from io import BytesIO
from typing import Any, List, Optional, cast

import openpyxl
import pandas as pd
from api.plans.views import GenericPlanViewSet
from api.utils.custom_permissions import IsAuthenticated, permission_required
from django.db.models import F, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from managers.models import Manager
from openpyxl import styles
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.viewsets import GenericViewSet

from .custom_schemas import *
from .serializers import PaymentReportFilterSerializer

logger = logging.getLogger()


class ExportPaymentReport(GenericPlanViewSet, GenericViewSet):

    @extend_schema(
        summary="Скачать отчет по выплатам",
        parameters=[PaymentReportFilterSerializer],
        responses=DEFAULT_FILE_RESPONSE,
    )
    @action(
        detail=False,
        methods=["get"],
        permission_classes=[IsAuthenticated],
        url_path="payment_report",
    )
    @permission_required("plans.export_payment_report")
    def payment_report(self, request: Request) -> HttpResponse:
        filter_serializer = PaymentReportFilterSerializer(data=request.query_params)
        filter_serializer.is_valid(raise_exception=True)

        start_date: Optional[datetime] = filter_serializer.validated_data["start_date"]
        end_date: Optional[datetime] = filter_serializer.validated_data["end_date"]
        managers: List[Manager] = filter_serializer.validated_data.get("managers", [])

        managers_qs = Manager.objects.filter(pk__in=[m.pk for m in managers])

        plans = self.filter_queryset(self.get_queryset())

        if not start_date:
            start_date = cast(datetime, plans.earliest("assigned_date").assigned_date)
        if not end_date:
            end_date = cast(datetime, plans.latest("assigned_date").assigned_date)
        if not managers:
            managers_qs = Manager.objects.all()
            managers = list(Manager.objects.all())

        plans = (
            plans.prefetch_related(
                Prefetch(
                    "manager_set",
                    queryset=managers_qs.prefetch_related("payment_registries"),
                )
            )
            .values(
                "assigned_date",
                "managers",
                "box_count",
                "client__name",
                "client__meta_client__name",
            )
            .filter(
                Q(managers__in=managers)
                & Q(managers__payment_registries__is_confirmed=True)
                & Q(managers__payment_registries__date=F("assigned_date"))
            )
            .annotate(
                payment_bonus=Coalesce(
                    Sum(
                        F("managers__payment_registries__payment")
                        + F("managers__payment_registries__bonus")
                    ),
                    Value(0),
                ),
                manager=F("managers__name"),
                manager_id=F("managers__id"),
                is_driver=F("managers__is_driver"),
                shop=F("client__name"),
                client=F("client__meta_client__name"),
                date=F("assigned_date"),
            )
            .values(
                "payment_bonus",
                "manager",
                "manager_id",
                "is_driver",
                "shop",
                "client",
                "date",
                "box_count",
            )
            .distinct()
        )

        buffer = generate_payment_report(list(plans))
        buffer.seek(0)

        # response = FileResponse(
        #     buffer,
        #     as_attachment=True,
        #     filename=f"ОТЧЕТ ПО ВЫПЛАТАМ {', '.join([m.name for m in managers_qs]) if 'manager' in filter_serializer.validated_data else ''} С {start_date.strftime('%d-%m-%Y')} ПО {end_date.strftime('%d-%m-%Y')}.xlsx",
        #     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        # )

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["Content-Disposition"] = (
            f"attachment; filename=\"ОТЧЕТ ПО ВЫПЛАТАМ {', '.join([m.name for m in managers_qs]) if 'manager' in filter_serializer.validated_data else ''} С {start_date.strftime('%d-%m-%Y')} ПО {end_date.strftime('%d-%m-%Y')}.xlsx\""
        )

        buffer.close()

        return response


def get_cols_by_letter(df):
    return {
        c: get_column_letter(cast(int, df.columns.get_loc(c)) + 1) for c in df.columns
    }


def generate_payment_report(table: list[dict[str, Any]]) -> BytesIO:
    # Empty table
    if len(table) == 0:
        raise ValueError(
            "Нет данных чтобы создать таблицу. Используйте другие фильтры или подтвердите больше выплат, и попробуйте занаво."
        )

    workbook = openpyxl.Workbook()
    # Formatting table
    # add styling
    if "header_cell" not in workbook.style_names:
        header_cell = styles.NamedStyle(name="header_cell")
        header_cell.alignment = styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_cell.border = styles.Border(
            left=styles.Side(style="thin"),
            right=styles.Side(style="thin"),
            top=styles.Side(style="thin"),
            bottom=styles.Side(style="thin"),
        )
        header_cell.font = styles.Font(name="Arial", size=12, bold=True)
        header_cell.number_format = "# ### ### ### ### ### ###"
        header_cell.fill = styles.PatternFill(
            start_color="efefef", end_color="efefef", fill_type="solid"
        )

        workbook.add_named_style(header_cell)
    if "generic_cell" not in workbook.style_names:
        generic_cell = styles.NamedStyle(name="generic_cell")
        generic_cell.alignment = styles.Alignment(vertical="center")
        generic_cell.border = styles.Border(
            left=styles.Side(style="thin"),
            right=styles.Side(style="thin"),
            top=styles.Side(style="thin"),
            bottom=styles.Side(style="thin"),
        )
        generic_cell.font = styles.Font(name="Tahoma", size=10)
        generic_cell.number_format = "# ### ### ### ### ### ###"
        workbook.add_named_style(generic_cell)
    if "subgroup_header_cell" not in workbook.style_names:
        subgroup_header_cell = styles.NamedStyle(name="subgroup_header_cell")
        subgroup_header_cell.alignment = styles.Alignment(vertical="center")
        subgroup_header_cell.border = styles.Border(
            left=styles.Side(style="thin"),
            right=styles.Side(style="thin"),
            top=styles.Side(style="thin"),
            bottom=styles.Side(style="thin"),
        )
        subgroup_header_cell.font = styles.Font(name="Tahoma", size=10, bold=True)
        subgroup_header_cell.number_format = "# ### ### ### ### ### ###"
        subgroup_header_cell.fill = styles.PatternFill(
            start_color="efefef", end_color="efefef", fill_type="solid"
        )

        workbook.add_named_style(subgroup_header_cell)
    if "footer_cell" not in workbook.style_names:
        footer_cell = styles.NamedStyle(name="footer_cell")
        footer_cell.alignment = styles.Alignment(vertical="center")
        footer_cell.border = styles.Border(
            left=styles.Side(style="thin"),
            right=styles.Side(style="thin"),
            top=styles.Side(style="thin"),
            bottom=styles.Side(style="thin"),
        )
        footer_cell.font = styles.Font(name="Tahoma", size=10, bold=True)
        footer_cell.number_format = "# ### ### ### ### ### ###"
        footer_cell.fill = styles.PatternFill(
            start_color="ffff00", end_color="ffff00", fill_type="solid"
        )

        workbook.add_named_style(footer_cell)

    total_visit_count = "SUMPRODUCT(({sheet}!{shop}:{shop}=INDEX({sheet}!{shop}:{shop}, {row}))*({sheet}!{manager}:{manager}=INDEX({sheet}!{manager}:{manager},{row})))"
    total_box_count = "SUMIFS({sheet}!{box_count}:{box_count}, {sheet}!{shop}:{shop}, INDEX({sheet}!{shop}:{shop}, {row}), {sheet}!{manager}:{manager}, INDEX({sheet}!{manager}:{manager}, {row}))"
    total_payment_bonus = {
        "driver": "SUMIFS({sheet}!{box_price}:{box_price}, {sheet}!{shop}:{shop}, INDEX({sheet}!{shop}:{shop}, {row}), {sheet}!{manager}:{manager}, INDEX({sheet}!{manager}:{manager}, {row}))",
        "manager": "SUMIFS({sheet}!{visit_price}:{visit_price}, {sheet}!{shop}:{shop}, INDEX({sheet}!{shop}:{shop}, {row}), {sheet}!{manager}:{manager}, INDEX({sheet}!{manager}:{manager}, {row}))",
    }

    box_price = "INDEX({sheet}!{payment_bonus}:{payment_bonus}, {row}) * INDEX({sheet}!{box_count}:{box_count}, {row}) / SUMIFS({sheet}!{box_count}:{box_count}, {sheet}!{date}:{date}, INDEX({sheet}!{date}:{date}, {row}), {sheet}!{manager}:{manager}, INDEX({sheet}!{manager}:{manager}, {row}))"
    visit_price = "INDEX({sheet}!{payment_bonus}:{payment_bonus}, {row}) / SUMPRODUCT(({sheet}!{date}:{date}=INDEX({sheet}!{date}:{date}, {row}))*({sheet}!{manager}:{manager}=INDEX({sheet}!{manager}:{manager},{row})))"

    df = pd.DataFrame(table)
    df["box_price"] = "=" + box_price.format(
        **get_cols_by_letter(df), sheet="raw_data", row="ROW()"
    )
    df["visit_price"] = "=" + visit_price.format(
        **get_cols_by_letter(df), sheet="raw_data", row="ROW()"
    )

    cols_by_letter = get_cols_by_letter(df)

    # Important to do it after cols_by_letter function, as it
    # errors when formatting of formulas are done.
    df["row"] = range(1, len(df) + 1)

    raw_data_ws = workbook.active or workbook.create_sheet("raw_data")
    raw_data_ws.title = "raw_data"
    for r in dataframe_to_rows(df, index=False, header=True):
        raw_data_ws.append(r)

    for manager, manager_group in df.groupby(by="manager"):
        manager = cast(str, manager)
        manager_ws = workbook.create_sheet("_" + manager)

        # https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
        manager_ws.sheet_properties.outlinePr.summaryBelow = False

        header = ["Клиент", "Коробки", "Кол-во выходов", "Выплаты", "Стоимость кор-ки"]
        manager_ws.append(header)
        for cell in list(manager_ws)[-1]:
            cell.style = "header_cell"

        group_rows_start = 2
        for client, client_manager_group in manager_group.groupby(by="client"):
            client_manager_group.drop_duplicates(["shop"], inplace=True)
            client_manager_group.sort_values(by=["shop"], inplace=True)

            # Sub-header, that contains aggregated data.
            client_manager_rows = client_manager_group["row"]
            client_body_df = pd.DataFrame(
                {
                    "shop": client_manager_group["shop"],
                    "total_box_count": client_manager_rows.apply(
                        lambda x: "="
                        + total_box_count.format(
                            **cols_by_letter, sheet="raw_data", row=x + 1
                        )
                    ),
                    "total_visit_count": client_manager_rows.apply(
                        lambda x: "="
                        + total_visit_count.format(
                            **cols_by_letter, sheet="raw_data", row=x + 1
                        )
                    ),
                    "total_payment_count": client_manager_rows.apply(
                        lambda x: "=ROUND({}, 0)".format(
                            (
                                total_payment_bonus["driver"]
                                if df.at[x - 1, "is_driver"]
                                else total_payment_bonus["manager"]
                            ).format(**cols_by_letter, sheet="raw_data", row=x + 1)
                        )
                    ),
                    "box_cost": client_manager_rows.apply(
                        lambda x: "=ROUND({}, 0)".format(
                            (
                                total_payment_bonus["driver"]
                                if df.at[x - 1, "is_driver"]
                                else total_payment_bonus["manager"]
                            ).format(**cols_by_letter, sheet="raw_data", row=x + 1)
                            + "/"
                            + total_box_count.format(
                                **cols_by_letter, sheet="raw_data", row=x + 1
                            )
                        )
                    ),
                }
            )
            body_size = len(client_body_df)
            body_range = (group_rows_start + 1, group_rows_start + body_size)
            client_header_df = pd.DataFrame(
                {
                    "shop": [client],
                    "total_box_count": ["=SUM(B{}:B{})".format(*body_range)],
                    "total_visit_count": ["=SUM(C{}:C{})".format(*body_range)],
                    "total_payment_count": [
                        "=ROUND(SUM(D{}:D{}), 0)".format(*body_range)
                    ],
                    "box_cost": ["=ROUND(D{0}/B{0}, 0)".format(group_rows_start)],
                }
            )

            for r in dataframe_to_rows(client_header_df, index=False, header=False):
                manager_ws.append(r)
                for cell in list(manager_ws)[-1]:
                    cell.style = "subgroup_header_cell"

            for r in dataframe_to_rows(client_body_df, index=False, header=False):
                manager_ws.append(r)
                for cell in list(manager_ws)[-1]:
                    cell.style = "generic_cell"

            manager_ws.row_dimensions.group(
                group_rows_start + 1,
                group_rows_start + len(client_body_df),
                hidden=True,
                outline_level=1,
            )

            # move to next sub-header row
            group_rows_start += len(client_header_df) + len(client_body_df)

        footer_range = (2, group_rows_start - 1)
        footer = [
            "Итого",
            "=SUM(B{}:B{})/2".format(*footer_range),
            "=SUM(C{}:C{})/2".format(*footer_range),
            "=ROUND(SUM(D{}:D{})/2, 0)".format(*footer_range),
            "=ROUND(D{0}/B{0}, 0)".format(group_rows_start),
        ]
        manager_ws.append(footer)
        for cell in list(manager_ws)[-1]:
            cell.style = "footer_cell"

        manager_ws.freeze_panes = "B2"

        # Change column width
        cols_width = [60, 25, 25, 25, 30]
        dim_holder = DimensionHolder(worksheet=manager_ws)
        for i, col in enumerate(
            range(manager_ws.min_column, manager_ws.max_column + 1)
        ):
            try:
                dim_holder[get_column_letter(col)] = ColumnDimension(
                    manager_ws, min=col, max=col, width=cols_width[i]
                )
            except Exception as _:
                continue
        manager_ws.column_dimensions = dim_holder

        # Change row height
        for i, _ in enumerate(
            manager_ws.iter_rows(min_row=1, max_row=group_rows_start)
        ):
            manager_ws.row_dimensions[i].height = 25

    workbook._sheets.sort(key=lambda ws: ws.title)

    buffer = BytesIO()
    workbook.save(buffer)

    return buffer
