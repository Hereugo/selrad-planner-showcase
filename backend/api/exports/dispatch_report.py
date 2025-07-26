import io
import logging
from datetime import datetime, timedelta
from enum import Enum
from typing import Optional, cast

import openpyxl
from api.plans.views import GenericPlanViewSet
from api.utils.custom_permissions import IsAuthenticated, permission_required
from django.db.models import QuerySet
from django.http import FileResponse, HttpResponse
from drf_spectacular.utils import extend_schema
from managers.models import Manager
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from openpyxl.utils import get_column_letter
from plans.models import Plan, PlanWorkItem
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.viewsets import GenericViewSet
from work_items.models import Shipment

from .custom_schemas import *
from .serializers import BaseFilterSerializer

logger = logging.getLogger()


class COL(Enum):
    DATE = 1
    CLIENT_NAME = 2
    TIME_SINCE_DISPATCH = 3
    DRIVERS = 4
    BOX_COUNT = 5
    TIME_SINCE_BOX_ARRIVAL = 6
    STATUS = 7
    MANAGER = 8
    COMMENT = 9


class ExportDispatchReport(GenericPlanViewSet, GenericViewSet):
    @extend_schema(
        summary="Cкачать диспетчерский отчет по периоду",
        description="",
        parameters=[BaseFilterSerializer],
        responses=DEFAULT_FILE_RESPONSE,
    )
    @action(
        detail=False,
        methods=["get"],
        permission_classes=[IsAuthenticated],
        url_path="dispatch_report",
    )
    @permission_required("plans.get_dispatch_report")
    def dispatch_report(self, request: Request) -> HttpResponse:
        """Cкачать диспетчерский отчет по периоду"""

        filter_serializer = BaseFilterSerializer(data=request.query_params)
        filter_serializer.is_valid(raise_exception=True)

        start_date: Optional[datetime] = filter_serializer.validated_data["start_date"]
        end_date: Optional[datetime] = filter_serializer.validated_data["end_date"]

        plans = self.filter_queryset(self.get_queryset())
        plans = plans.order_by("assigned_date")

        if not start_date:
            start_date = cast(datetime, plans.earliest("assigned_date").assigned_date)
        if not end_date:
            end_date = cast(datetime, plans.latest("assigned_date").assigned_date)

        work_items_shipments = (
            PlanWorkItem.objects.filter(plan__in=plans, content_type__model="shipment")
            .prefetch_related("content_object")
            .select_related("plan")
        )

        buffer = generate_dispatch_report(work_items_shipments, start_date, end_date)
        buffer.seek(0)

        # response = FileResponse(
        #     buffer,
        #     as_attachment=True,
        #     filename=f"ОТЧЕТ ПО ДИСПЕЧЕРСКОМУ С {start_date.strftime('%d-%m-%Y')} ПО {end_date.strftime('%d-%m-%Y')}.xlsx",
        #     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        # )

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["Content-Disposition"] = (
            f"attachment; filename=\"ОТЧЕТ ПО ДИСПЕЧЕРСКОМУ С {start_date.strftime('%d-%m-%Y')} ПО {end_date.strftime('%d-%m-%Y')}.xlsx\""
        )

        buffer.close()

        return response


def generate_dispatch_report(
    work_items_shipments: QuerySet[PlanWorkItem],
    start_date: datetime,
    end_date: datetime,
):
    workbook = openpyxl.load_workbook("./static/docs/standard_dispatch_report.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    ws.cell(row=4, column=1).value = (
        f"Период: {start_date.strftime('%d/%m/%Y')} по {end_date.strftime('%d/%m/%Y')}"
    )

    row = 7
    for i, work_item in enumerate(work_items_shipments):
        shipment: Shipment = work_item.content_object
        plan: Plan = work_item.plan
        drivers: QuerySet[Manager] = plan.managers.filter(is_driver=True)

        for col in range(1, len(COL) + 1):
            ws.cell(row=row + i, column=col).style = "general_style"

        ws.cell(row=row + i, column=COL.DATE.value).value = plan.assigned_date.strftime(
            "%d/%m/%Y"
        )
        ws.cell(row=row + i, column=COL.CLIENT_NAME.value).value = plan.client.name
        ws.cell(row=row + i, column=COL.TIME_SINCE_DISPATCH.value).value = (
            "-"
            if plan.time_since_first_dispatch is None
            else (plan.time_since_first_dispatch + timedelta(hours=5)).strftime("%H:%M")
        )
        ws.cell(row=row + i, column=COL.DRIVERS.value).value = ", ".join(
            [d.name for d in drivers.all()]
        )
        ws.cell(row=row + i, column=COL.BOX_COUNT.value).value = shipment.box_count
        ws.cell(row=row + i, column=COL.TIME_SINCE_BOX_ARRIVAL.value).value = (
            "-"
            if shipment.time_since_last_box_arrival is None
            else (shipment.time_since_last_box_arrival + timedelta(hours=5)).strftime(
                "%H:%M"
            )
        )
        ws.cell(row=row + i, column=COL.STATUS.value).value = dict(
            Shipment.CHOICES
        ).get(shipment.status, "")
        ws.cell(row=row + i, column=COL.MANAGER.value).value = (
            "" if shipment.completed_by is None else shipment.completed_by.name
        )
        ws.cell(row=row + i, column=COL.COMMENT.value).value = shipment.comment

    last_row = row + len(work_items_shipments)

    ws.cell(row=last_row, column=COL.DATE.value).value = "ИТОГО:"
    ws.cell(row=last_row, column=COL.DATE.value).font = Font(bold=True)
    ws.cell(row=last_row, column=COL.BOX_COUNT.value).value = (
        "=SUM({0}{1}:{0}{2})".format(
            get_column_letter(COL.BOX_COUNT.value), 7, last_row - 1
        )
    )
    ws.cell(row=last_row, column=COL.BOX_COUNT.value).font = Font(bold=True)
    ws.cell(row=last_row, column=COL.BOX_COUNT.value).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer
