from io import BytesIO
from datetime import datetime

from django.db.models import QuerySet

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side

from plans.models import Plan
from managers.models import Manager


def generate_dispatch_list(
    plans: QuerySet[Plan],
    manager: Manager,
    comment: str,
    start_date: datetime,
    end_date: datetime,
):
    workbook = openpyxl.load_workbook("./static/docs/standard_dispatch_list.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment.wrap_text = True

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    # Setup title:
    ws.cell(row=1, column=1).value = f"Диспетчерсктй лист {manager}"

    # Listing all parameters:
    ws.cell(row=2, column=1).value = "Параметры:"
    ws.cell(row=3, column=1).value = (
        f"Период: {start_date.strftime('%d-%m-%Y')} с {end_date.strftime('%d-%m-%Y')}"
    )
    ws.cell(row=4, column=1).value = f"Менеджер: {manager}"

    # Setup table data:
    row_offset: int = 7
    for i, plan in enumerate(plans, start=row_offset):
        ws.cell(row=i, column=1).value = i - row_offset + 1
        ws.cell(row=i, column=2).value = plan.client.name
        ws.cell(row=i, column=3).value = plan.box_count
        ws.cell(row=i, column=4).value = plan.client.address.street
        ws.cell(row=i, column=5).value = ", ".join(
            [str(m) for m in plan.managers.all()]
        )
        ws.cell(row=i, column=7).value = plan.comment

    ws.cell(row=len(plans) + row_offset + 2, column=1).value = comment

    buffer = BytesIO()
    workbook.save(buffer)

    return buffer
