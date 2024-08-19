import io
import logging
from itertools import groupby

import openpyxl
from openpyxl.styles import Border, Side, NamedStyle, Font, PatternFill, Alignment


logger = logging.getLogger(__name__)


COL_DICT = {
    "client": 1,
    "address": 2,
    "manager": 3,
    "work_items": 4,
    "comment": 5,
    "shipment_cost": 6,
    "box_count": 7,
}


def ru_week_name(date):
    week_day = date.weekday()
    if week_day == 0:
        return "ПОНЕДЕЛЬНИК"
    elif week_day == 1:
        return "ВТОРНИК"
    elif week_day == 2:
        return "СРЕДА"
    elif week_day == 3:
        return "ЧЕТВЕРГ"
    elif week_day == 4:
        return "ПЯТНИЦА"
    elif week_day == 5:
        return "СУББОТА"
    elif week_day == 6:
        return "ВОСКРЕСЕНЬЕ"
    else:
        return "Неверный день недели"


def gen_header(ws, row, title, sc, ec):
    ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
    ws.cell(row=row, column=1).value = title
    ws.cell(row=row, column=1).style = "head_cell"


def generate_excelsheet_by_plan(plans, earliest_date, latest_date):
    workbook = openpyxl.load_workbook("./static/docs/standard_plan.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "head_cell" not in workbook.style_names:
        head_style = NamedStyle(name="head_cell")
        head_style.font = Font(color="000000", size=10)
        head_style.fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        workbook.add_named_style(head_style)

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment.wrap_text = True
        general_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0"

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    gen_header(
        ws,
        1,
        f"ПЛАНЫ С {earliest_date.strftime('%d.%m.%Y')} ПО {latest_date.strftime('%d.%m.%Y')}",
        1,
        5,
    )

    row = 3
    for assigned_date, plans_by_day in groupby(plans, key=lambda p: p.assigned_date):
        plans_by_day = sorted(
            plans_by_day,
            key=lambda p: (p.client.is_hidden_on_map, p.created_at),
        )

        if len(plans_by_day) == 0:
            continue

        gen_header(
            ws,
            row,
            f"{assigned_date.strftime('%d.%m')} - {ru_week_name(assigned_date)}",
            1,
            5,
        )

        row += 2
        plan_count = 0
        for i, plan in enumerate(plans_by_day):
            for col in range(1, len(COL_DICT) + 1):
                ws.cell(row=row + i, column=col).style = "general_style"

            ws.cell(row=row + i, column=COL_DICT["client"]).value = (
                f"{i + 1}. {plan.client.name}"
            )
            ws.cell(row=row + i, column=COL_DICT["address"]).value = (
                plan.client.address.street
            )
            ws.cell(row=row + i, column=COL_DICT["manager"]).value = ", ".join(
                [str(m) for m in plan.managers.all()]
            )
            ws.cell(row=row + i, column=COL_DICT["work_items"]).value = ", ".join(
                [str(w) for w in plan.work_items.all()]
            )
            ws.cell(row=row + i, column=COL_DICT["comment"]).value = plan.comment
            ws.cell(row=row + i, column=COL_DICT["shipment_cost"]).value = (
                plan.shipment_cost()
            )
            ws.cell(row=row + i, column=COL_DICT["shipment_cost"]).alignment = (
                Alignment(horizontal="right")
            )
            ws.cell(row=row + i, column=COL_DICT["box_count"]).value = plan.box_count
            ws.cell(row=row + i, column=COL_DICT["box_count"]).alignment = Alignment(
                horizontal="right"
            )

            plan_count += 1

        row += plan_count + 1

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer


COL_DICT_REPORT = {
    "client": 1,
    "address": 2,
    "work_items": 3,
    "comment": 4,
    "shipment_cost": 5,
    "box_count": 6,
}


def generate_excelsheet_by_manager(plans, manager, earliest_date, latest_date):
    workbook = openpyxl.load_workbook("./static/docs/standard_report.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "head_cell" not in workbook.style_names:
        head_style = NamedStyle(name="head_cell")
        head_style.font = Font(color="000000", size=10)
        head_style.fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        workbook.add_named_style(head_style)

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment.wrap_text = True
        general_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0"

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    gen_header(
        ws,
        1,
        f"ОТЧЕТ {manager} С {earliest_date.strftime('%d.%m.%Y')} ПО {latest_date.strftime('%d.%m.%Y')}",
        1,
        4,
    )

    row = 3
    for assigned_date, plans_by_day in groupby(plans, key=lambda p: p.assigned_date):
        plans_by_day = sorted(
            plans_by_day,
            key=lambda p: (p.client.is_hidden_on_map, p.created_at),
        )

        if len(plans_by_day) == 0:
            continue

        gen_header(
            ws,
            row,
            f"{assigned_date.strftime('%d.%m')} - {ru_week_name(assigned_date)}",
            1,
            4,
        )
        row += 2
        plan_count = 0
        for i, plan in enumerate(plans_by_day):
            for col in range(1, len(COL_DICT_REPORT) + 1):
                ws.cell(row=row + i, column=col).style = "general_style"

            ws.cell(row=row + i, column=COL_DICT_REPORT["client"]).value = (
                f"{i + 1}. {plan.client.name}"
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["address"]).value = (
                plan.client.address.street
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["work_items"]).value = (
                ", ".join([str(w) for w in plan.work_items.all()])
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["comment"]).value = plan.comment
            ws.cell(row=row + i, column=COL_DICT_REPORT["shipment_cost"]).value = (
                plan.shipment_cost()
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["shipment_cost"]).alignment = (
                Alignment(horizontal="right")
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["box_count"]).value = (
                plan.box_count
            )
            ws.cell(row=row + i, column=COL_DICT_REPORT["box_count"]).alignment = (
                Alignment(horizontal="right")
            )

            plan_count += 1

        row += plan_count + 1

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer
