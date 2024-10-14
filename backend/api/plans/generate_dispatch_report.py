import io
import logging
from enum import Enum
from datetime import datetime, timedelta

from django.db.models import QuerySet
import openpyxl
from openpyxl.styles import Border, Font, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter

from plans.models import Plan, PlanWorkItem
from work_items.models import Shipment
from managers.models import Manager

logger = logging.getLogger(__name__)


class COL(Enum):
    DATE = 1
    CLIENT_NAME = 2
    TIME_SINCE_DISPATCH = 3
    DRIVERS = 4
    BOX_COUNT = 5
    TIME_SINCE_BOX_ARRIVAL = 6
    STATUS = 7
    MANAGER = 8
    COMMENT = 9


def generate_dispatch_report(
    work_items_shipments: QuerySet[PlanWorkItem],
    start_date: datetime,
    end_date: datetime,
):
    workbook = openpyxl.load_workbook("./static/docs/standard_dispatch_report.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    ws.cell(row=4, column=1).value = (
        f"Период: {start_date.strftime('%d/%m/%Y')} по {end_date.strftime('%d/%m/%Y')}"
    )

    row = 7
    for i, work_item in enumerate(work_items_shipments):
        shipment: Shipment = work_item.content_object
        plan: Plan = work_item.plan
        drivers: QuerySet[Manager] = plan.managers.filter(is_driver=True)

        for col in range(1, len(COL) + 1):
            ws.cell(row=row + i, column=col).style = "general_style"

        ws.cell(row=row + i, column=COL.DATE.value).value = plan.assigned_date.strftime(
            "%d/%m/%Y"
        )
        ws.cell(row=row + i, column=COL.CLIENT_NAME.value).value = plan.client.name
        ws.cell(row=row + i, column=COL.TIME_SINCE_DISPATCH.value).value = (
            "-"
            if plan.time_since_last_dispatch is None
            else (plan.time_since_last_dispatch + timedelta(hours=5)).strftime("%H:%M")
        )
        ws.cell(row=row + i, column=COL.DRIVERS.value).value = ", ".join(
            [d.name for d in drivers.all()]
        )
        ws.cell(row=row + i, column=COL.BOX_COUNT.value).value = shipment.box_count
        ws.cell(row=row + i, column=COL.TIME_SINCE_BOX_ARRIVAL.value).value = (
            "-"
            if shipment.time_since_last_box_arrival is None
            else (shipment.time_since_last_box_arrival + timedelta(hours=5)).strftime(
                "%H:%M"
            )
        )
        ws.cell(row=row + i, column=COL.STATUS.value).value = dict(
            Shipment.CHOICES
        ).get(shipment.status, "")
        ws.cell(row=row + i, column=COL.MANAGER.value).value = (
            "" if shipment.completed_by is None else shipment.completed_by.name
        )
        ws.cell(row=row + i, column=COL.COMMENT.value).value = shipment.comment

    last_row = row + len(work_items_shipments) + 1

    ws.cell(row=last_row, column=COL.DATE.value).value = "ИТОГО:"
    ws.cell(row=last_row, column=COL.DATE.value).font = Font(bold=True)
    ws.cell(row=last_row, column=COL.BOX_COUNT.value).value = (
        "SUM({0}{1}:{0}{2})".format(
            get_column_letter(COL.BOX_COUNT.value), 7, last_row - 1
        )
    )
    ws.cell(row=last_row, column=COL.BOX_COUNT.value).font = Font(bold=True)

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer
