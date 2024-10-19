import io
import logging

from django.contrib.postgres.aggregates import StringAgg
from django.db.models import Sum

import openpyxl
from openpyxl.styles import Border, Side, NamedStyle, Font, PatternFill

from .custom_filters import PlanFilter
from clients.models import Client
from plans.models import WorkItem
from managers.models import Manager


logger = logging.getLogger(__name__)


def gen_header(ws, start_row, **kwargs):
    period_1_start = kwargs["period_1"]["start"].strftime("%d-%m-%Y")
    period_1_end = kwargs["period_1"]["end"].strftime("%d-%m-%Y")
    period_2_start = kwargs["period_2"]["start"].strftime("%d-%m-%Y")
    period_2_end = kwargs["period_2"]["end"].strftime("%d-%m-%Y")
    managers = kwargs.get("managers", [])
    work_items = kwargs.get("work_items", [])

    title = f"СРАВНИТЬ С {period_1_start} ПО {period_1_end} ПРОТИВ {period_2_start} ПО {period_2_end}"

    ws.cell(row=start_row, column=1).value = title
    ws.cell(row=start_row, column=1).style = "head_cell"
    ws.cell(row=start_row + 2, column=1).value = (
        f"Прошлый период: {period_1_start} по {period_1_end}"
    )
    ws.cell(row=start_row + 3, column=1).value = (
        f"Текущий период: {period_2_start} по {period_2_end}"
    )
    ws.cell(row=start_row + 4, column=1).value = f"Менеджеры: " + ", ".join(
        [m.name for m in managers]
    )
    ws.cell(row=start_row + 5, column=1).value = f"Работы: " + ", ".join(
        [w.name for w in work_items]
    )


def generate_compare_years(period_1, period_2, params):
    workbook = openpyxl.load_workbook("./static/docs/standard_compare_years.xlsx")
    ws = workbook.active or workbook.create_sheet("Sheet1")

    if "head_cell" not in workbook.style_names:
        head_style = NamedStyle(name="head_cell")
        head_style.font = Font(color="000000", size=12, bold=True)
        workbook.add_named_style(head_style)

    if "headers_style" not in workbook.style_names:
        headers_style = NamedStyle(name="headers_style")
        headers_style.alignment.wrap_text = True
        headers_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0"
        headers_style.font = Font(color="000000", bold=True)
        headers_style.fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        headers_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(headers_style)

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment.wrap_text = True
        general_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0"

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    managers = Manager.objects.filter(id__in=params.get("managers", []))
    work_items = WorkItem.objects.filter(id__in=params.get("work_items", []))
    gen_header(
        ws,
        1,
        period_1=period_1,
        period_2=period_2,
        managers=managers,
        work_items=work_items,
    )

    table = {}
    for client in Client.objects.all():
        meta_client_name = (
            client.meta_client.name if client.meta_client else "НЕНАЗНАЧЕН"
        )
        if not meta_client_name in table:
            table[meta_client_name] = []

        period_1_plans = client.plans.filter(
            assigned_date__gte=period_1["start"],
            assigned_date__lte=period_1["end"],
        )
        period_2_plans = client.plans.filter(
            assigned_date__gte=period_2["start"],
            assigned_date__lte=period_2["end"],
        )
        if len(managers):
            period_1_plans = period_1_plans.filter(managers__in=managers)
            period_2_plans = period_2_plans.filter(managers__in=managers)
        if len(work_items):
            period_1_plans = period_1_plans.filter(work_items__in=work_items)
            period_2_plans = period_2_plans.filter(work_items__in=work_items)

        if not period_1_plans.exists() and not period_2_plans.exists():
            logger.debug(
                f"Skipping client {client} as no plans were given for it in both periods"
            )
            continue

        logger.debug(period_1_plans)
        logger.debug(period_2_plans)
        logger.debug(params)

        table[meta_client_name].append(
            {
                "client_name": client.name,
                "period_1": {
                    "total_box_count": period_1_plans.aggregate(Sum("box_count"))[
                        "box_count__sum"
                    ],
                    "total_shipment_cost": period_1_plans.aggregate(
                        x=StringAgg("shipment_cost_formula", delimiter="+")
                    )["x"]
                    or "0",
                },
                "period_2": {
                    "total_box_count": period_2_plans.aggregate(Sum("box_count"))[
                        "box_count__sum"
                    ],
                    "total_shipment_cost": period_2_plans.aggregate(
                        x=StringAgg("shipment_cost_formula", delimiter="+")
                    )["x"]
                    or "0",
                },
            }
        )

    # Grouping rows
    # One issue that incountered was that groups merge if they overlap and on the same outline level
    # https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 11
    f = []
    for meta_client_name, clients in table.items():
        if len(clients) == 0:
            continue

        num_clients = len(clients)
        ws.cell(row=row, column=1).value = meta_client_name
        # table 1
        ws.cell(row=row, column=2).value = "=SUM(B{}:B{})".format(
            row + 1, row + num_clients
        )
        ws.cell(row=row, column=3).value = "=SUM(C{}:C{})".format(
            row + 1, row + num_clients
        )
        # table 2
        ws.cell(row=row, column=5).value = "=SUM(E{}:E{})".format(
            row + 1, row + num_clients
        )
        ws.cell(row=row, column=6).value = "=SUM(F{}:F{})".format(
            row + 1, row + num_clients
        )
        # table 3
        ws.cell(row=row, column=8).value = "=E{0}/B{0}".format(row)
        ws.cell(row=row, column=9).value = "=F{0}/C{0}".format(row)

        for col in [1, 2, 3, 5, 6, 8, 9]:
            ws.cell(row=row, column=col).style = "headers_style"

        for offset, client in enumerate(clients, 1):
            # table 1
            ws.cell(row=row + offset, column=1).value = client["client_name"]
            ws.cell(row=row + offset, column=2).value = client["period_1"][
                "total_box_count"
            ]
            ws.cell(row=row + offset, column=3).value = client["period_2"][
                "total_box_count"
            ]
            # table 2
            ws.cell(row=row + offset, column=5).value = (
                "=" + client["period_1"]["total_shipment_cost"]
            )
            ws.cell(row=row + offset, column=6).value = (
                "=" + client["period_2"]["total_shipment_cost"]
            )
            # table 3
            ws.cell(row=row + offset, column=8).value = "=E{0}/B{0}".format(
                row + offset
            )
            ws.cell(row=row + offset, column=9).value = "=F{0}/C{0}".format(
                row + offset
            )

            for col in [1, 2, 3, 5, 6, 8, 9]:
                ws.cell(row=row + offset, column=col).style = "general_style"

        ws.row_dimensions.group(
            row + 1,
            row + num_clients,
            hidden=True,
            outline_level=1,
        )

        f.append(row)
        row += num_clients + 1

    ws.cell(row=row, column=1).value = "ИТОГО:"
    ws.cell(row=row, column=2).value = "=" + "+".join([f"B{x}" for x in f])
    ws.cell(row=row, column=3).value = "=" + "+".join([f"C{x}" for x in f])
    ws.cell(row=row, column=5).value = "=" + "+".join([f"E{x}" for x in f])
    ws.cell(row=row, column=6).value = "=" + "+".join([f"F{x}" for x in f])
    ws.cell(row=row, column=8).value = "=E{0}/B{0}".format(row)
    ws.cell(row=row, column=9).value = "=F{0}/C{0}".format(row)
    for col in [1, 2, 3, 5, 6, 8, 9]:
        ws.cell(row=row, column=col).style = "headers_style"
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

    buffer = io.BytesIO()
    workbook.save(buffer)

    return buffer
