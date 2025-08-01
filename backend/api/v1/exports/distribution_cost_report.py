import logging
from datetime import datetime
from io import BytesIO
from typing import Any, List, Optional, cast

import openpyxl
import pandas as pd
from django.db.models import F, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce, Concat
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.viewsets import GenericViewSet

from api.v1.exports.custom_schemas import *
from api.v1.exports.serializers import DistributionCostReportFilterSerializer
from api.v1.plans.views import GenericPlanViewSet
from api.v1.utils.custom_permissions import IsAuthenticated, permission_required
from clients.models import Client
from managers.models import Manager

logger = logging.getLogger(__name__)


class ExportDistributionCostReport(GenericPlanViewSet, GenericViewSet):

    @extend_schema(
        summary="Скачать отчет по стоимоти дистрибуции",
        parameters=[DistributionCostReportFilterSerializer],
        responses=DEFAULT_FILE_RESPONSE,
    )
    @action(
        detail=False,
        methods=["get"],
        permission_classes=[IsAuthenticated],
        url_path="distribution_cost_report",
    )
    @permission_required("plans.export_distribution_cost_report")
    def distribution_cost_report(self, request: Request) -> HttpResponse:
        filter_serializer = DistributionCostReportFilterSerializer(
            data=request.query_params
        )
        filter_serializer.is_valid(raise_exception=True)

        start_date: Optional[datetime] = filter_serializer.validated_data["start_date"]
        end_date: Optional[datetime] = filter_serializer.validated_data["end_date"]
        managers: List[Manager] = filter_serializer.validated_data.get("managers", [])

        managers_qs = Manager.objects.filter(pk__in=[m.pk for m in managers])

        plans = self.filter_queryset(self.get_queryset())

        if not start_date:
            start_date = cast(datetime, plans.earliest("assigned_date").assigned_date)
        if not end_date:
            end_date = cast(datetime, plans.latest("assigned_date").assigned_date)
        if not managers:
            managers_qs = Manager.objects.all()
            managers = list(Manager.objects.all())

        plans = (
            plans.prefetch_related(
                Prefetch(
                    "manager_set",
                    queryset=managers_qs.prefetch_related("payment_registries"),
                )
            )
            .values(
                "assigned_date",
                "managers",
                "box_count",
                "shipment_cost_formula",
                "client__name",
                "client__meta_client__name",
            )
            .filter(
                Q(managers__in=managers)
                & Q(managers__payment_registries__is_confirmed=True)
                & Q(managers__payment_registries__date=F("assigned_date"))
            )
            .annotate(
                payment_bonus=Coalesce(
                    Sum(
                        F("managers__payment_registries__payment")
                        + F("managers__payment_registries__bonus")
                    ),
                    Value(0),
                ),
                manager=F("managers__name"),
                manager_id=F("managers__id"),
                is_driver=F("managers__is_driver"),
                shop=F("client__name"),
                client=F("client__meta_client__name"),
                date=F("assigned_date"),
                shipment_cost=Concat(Value("="), F("shipment_cost_formula")),
            )
            .values(
                "payment_bonus",
                "manager",
                "manager_id",
                "is_driver",
                "shop",
                "client",
                "date",
                "box_count",
                "shipment_cost",
            )
            .distinct()
        )

        buffer = generate_distribution_cost_report(list(plans))
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["Content-Disposition"] = (
            f"attachment; filename=\"ОТЧЕТ ПО СТОИМОТИ ДИСТРИБУЦИИ {', '.join([m.name for m in managers_qs]) if 'manager' in filter_serializer.validated_data else ''} С {start_date.strftime('%d-%m-%Y')} ПО {end_date.strftime('%d-%m-%Y')}.xlsx\""
        )

        buffer.close()

        return response


def get_cols_by_letter(df):
    return {
        c: get_column_letter(cast(int, df.columns.get_loc(c)) + 1) for c in df.columns
    }


def generate_distribution_cost_report(table: list[dict[str, Any]]) -> BytesIO:
    workbook = openpyxl.Workbook()

    if "headers_style" not in workbook.style_names:
        headers_style = NamedStyle(name="headers_style")
        headers_style.font = Font(color="000000", bold=True, name="Arial")
        headers_style.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        headers_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(headers_style)

    if "subheaders_style" not in workbook.style_names:
        subheaders_style = NamedStyle(name="subheaders_style")
        subheaders_style.alignment.wrap_text = True
        subheaders_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0;-### ### ### ### ### ### ### ### ### ### ##0;"
        subheaders_style.font = Font(color="000000", bold=True, name="Arial")
        subheaders_style.fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        subheaders_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(subheaders_style)

    if "general_style" not in workbook.style_names:
        general_style = NamedStyle(name="general_style")
        general_style.alignment.wrap_text = True
        general_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0;-### ### ### ### ### ### ### ### ### ### ##0;"
        general_style.font = Font(color="000000", name="Arial")

        general_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(general_style)

    if "footer_style" not in workbook.style_names:
        footer_style = NamedStyle(name="footer_style")
        footer_style.font = Font(color="000000", bold=True, name="Arial")
        footer_style.alignment.wrap_text = True
        footer_style.number_format = "### ### ### ### ### ### ### ### ### ### ##0;-### ### ### ### ### ### ### ### ### ### ##0;"
        footer_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        workbook.add_named_style(footer_style)

    # add raw table in separate spreadsheet
    total_box_count = "SUM(MAP(UNIQUE(FILTER({sheet}!{date}:{date}, {sheet}!{shop}:{shop}={sheet}!{shop}{row})), LAMBDA(h, XLOOKUP(h, FILTER({sheet}!{date}:{date}, {sheet}!{shop}:{shop}={sheet}!{shop}{row}), FILTER({sheet}!{box_count}:{box_count}, {sheet}!{shop}:{shop}={sheet}!{shop}{row})))))"
    total_shipment_cost = "SUM(MAP(UNIQUE(FILTER({sheet}!{date}:{date}, {sheet}!{shop}:{shop}={sheet}!{shop}{row})), LAMBDA(h, XLOOKUP(h, FILTER({sheet}!{date}:{date}, {sheet}!{shop}:{shop}={sheet}!{shop}{row}), FILTER({sheet}!{shipment_cost}:{shipment_cost}, {sheet}!{shop}:{shop}={sheet}!{shop}{row})))))"

    total_payment_bonus = {
        "driver": 'ROUND(SUMIFS({sheet}!{box_price}:{box_price}, {sheet}!{shop}:{shop}, INDEX({sheet}!{shop}:{shop}, {row}), {sheet}!{manager}:{manager}, "{manager_name}"), 0)',
        "manager": 'ROUND(SUMIFS({sheet}!{visit_price}:{visit_price}, {sheet}!{shop}:{shop}, INDEX({sheet}!{shop}:{shop}, {row}), {sheet}!{manager}:{manager}, "{manager_name}"), 0)',
    }

    box_price = "INDEX({sheet}!{payment_bonus}:{payment_bonus}, {row}) * INDEX({sheet}!{box_count}:{box_count}, {row}) / SUMIFS({sheet}!{box_count}:{box_count}, {sheet}!{date}:{date}, INDEX({sheet}!{date}:{date}, {row}), {sheet}!{manager}:{manager}, INDEX({sheet}!{manager}:{manager}, {row}))"
    visit_price = "INDEX({sheet}!{payment_bonus}:{payment_bonus}, {row}) / SUMPRODUCT(({sheet}!{date}:{date}=INDEX({sheet}!{date}:{date}, {row}))*({sheet}!{manager}:{manager}=INDEX({sheet}!{manager}:{manager},{row})))"

    df = pd.DataFrame(table)
    df["box_price"] = "=" + box_price.format(
        **get_cols_by_letter(df), sheet="raw_data", row="ROW()"
    )
    df["visit_price"] = "=" + visit_price.format(
        **get_cols_by_letter(df), sheet="raw_data", row="ROW()"
    )

    drivers = df[df["is_driver"]]["manager"].unique()
    managers = df[~df["is_driver"]]["manager"].unique()

    cols_by_letter = get_cols_by_letter(df)

    # Important to do it after cols_by_letter function, as it
    # errors when formatting of formulas are done.
    df["row"] = range(1, len(df) + 1)

    # create and fillout main worksheet
    ws = workbook.active or workbook.create_sheet()

    raw_data_ws = workbook.create_sheet("raw_data")
    raw_data_ws.title = "raw_data"
    for r in dataframe_to_rows(df, index=False, header=True):
        raw_data_ws.append(r)
    raw_data_ws.sheet_state = "hidden"

    # setup header cells for table1
    ws.cell(row=1, column=1).value = "магазины / клиенты"
    ws.cell(row=1, column=2).value = "Кол-во коробок (кор)"
    ws.cell(row=1, column=3).value = "Сумма отгрузки (₸)"
    for i in range(1, 3 + 1):
        ws.cell(row=1, column=i).style = "headers_style"

    table2_offset = 4
    for i, driver in enumerate(
        list(drivers) + ["ВОДИТЕЛИ", "СТОИМОСТЬ КОРОБКИ", "ПРОЦЕНТ"], start=1
    ):
        ws.cell(row=1, column=table2_offset + i).value = driver
        ws.cell(row=1, column=table2_offset + i).style = "headers_style"

    table3_offset = table2_offset + len(drivers) + 3 + 1
    for i, manager in enumerate(
        list(managers) + ["ДЕВОЧКИ", "СТОИМОСТЬ КОРОБКИ", "ПРОЦЕНТ"], start=1
    ):
        ws.cell(row=1, column=table3_offset + i).value = manager
        ws.cell(row=1, column=table3_offset + i).style = "headers_style"

    # https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
    ws.sheet_properties.outlinePr.summaryBelow = False

    group_rows_start = 2
    for client, shop_group in df.groupby(by="client"):

        shop_group.drop_duplicates(["shop"], inplace=True)
        shop_group.sort_values(by=["shop"], inplace=True)

        # TABLE 1 CREATION
        table1_df = pd.DataFrame(
            {
                "shop": shop_group["shop"],
                "total_box_count": shop_group["row"].apply(
                    lambda x: "="
                    + total_box_count.format(
                        **cols_by_letter, sheet="raw_data", row=x + 1
                    )
                ),
                "total_shipment_cost": shop_group["row"].apply(
                    lambda x: "="
                    + total_shipment_cost.format(
                        **cols_by_letter, sheet="raw_data", row=x + 1
                    )
                ),
            }
        )
        table1_size = len(table1_df)
        table1_range = (group_rows_start + 1, group_rows_start + table1_size)
        table1_header_df = pd.DataFrame(
            {
                "shop": [client],
                "total_box_count": [
                    "=SUM({box_count_col}{}:{box_count_col}{})".format(
                        *table1_range,
                        box_count_col="B",
                    )
                ],
                "total_shipment_cost": [
                    "=SUM({shipment_cost_col}{}:{shipment_cost_col}{})".format(
                        *table1_range,
                        shipment_cost_col="C",
                    )
                ],
            }
        )
        series_range = pd.Series(
            range(group_rows_start + 1, group_rows_start + table1_size + 1)
        )

        table2_df = pd.DataFrame(
            {
                driver: shop_group["row"].apply(
                    lambda x: "="
                    + total_payment_bonus["driver"].format(
                        **cols_by_letter,
                        sheet="raw_data",
                        manager_name=driver,
                        row=x + 1,
                    )
                )
                for driver in drivers
            }
        )
        stat_table2_df = pd.DataFrame(
            {
                "total_payment": series_range.apply(
                    lambda x: "=SUM({}{row}:{}{row})".format(
                        get_column_letter(table2_offset + 1),
                        get_column_letter(table2_offset + len(drivers)),
                        row=x,
                    )
                ),
                "box_cost": series_range.apply(
                    lambda x: "={}{row}/{box_count_col}{row}".format(
                        get_column_letter(table2_offset + len(drivers) + 1),
                        box_count_col="B",
                        row=x,
                    )
                ),
                "percent": series_range.apply(
                    lambda x: "={}{row}/{shipment_cost_col}{row}".format(
                        get_column_letter(table2_offset + len(drivers) + 1),
                        shipment_cost_col="C",
                        row=x,
                    )
                ),
            }
        )

        table2_df = pd.concat(
            [table2_df.reset_index(drop=True), stat_table2_df.reset_index(drop=True)],
            axis=1,
        )

        table3_df = pd.DataFrame(
            {
                **{
                    manager: shop_group["row"].apply(
                        lambda x: "="
                        + total_payment_bonus["manager"].format(
                            **cols_by_letter,
                            sheet="raw_data",
                            manager_name=manager,
                            row=x + 1,
                        )
                    )
                    for manager in managers
                },
            }
        )

        stat_table3_df = pd.DataFrame(
            {
                "total_payment": series_range.apply(
                    lambda x: "=SUM({}{row}:{}{row})".format(
                        get_column_letter(table3_offset + 1),
                        get_column_letter(table3_offset + len(managers)),
                        row=x,
                    )
                ),
                "box_cost": series_range.apply(
                    lambda x: "={}{row}/{box_count_col}{row}".format(
                        get_column_letter(table3_offset + len(managers) + 1),
                        box_count_col="B",
                        row=x,
                    )
                ),
                "percent": series_range.apply(
                    lambda x: "={}{row}/{shipment_cost_col}{row}".format(
                        get_column_letter(table3_offset + len(managers) + 1),
                        shipment_cost_col="C",
                        row=x,
                    )
                ),
            }
        )

        table3_df = pd.concat(
            [table3_df.reset_index(drop=True), stat_table3_df.reset_index(drop=True)],
            axis=1,
        )

        table2_header_df = pd.DataFrame(
            {
                **{
                    driver: [
                        "=SUM({manager_col}{}:{manager_col}{})".format(
                            *table1_range,
                            manager_col=get_column_letter(i),
                        )
                    ]
                    for i, driver in enumerate(drivers, start=table2_offset + 1)
                },
                "total_payment": [
                    "=SUM({}{row}:{}{row})".format(
                        get_column_letter(table2_offset + 1),
                        get_column_letter(table2_offset + len(drivers)),
                        row=group_rows_start,
                    )
                ],
                "box_cost": [
                    "={}{row}/{box_count_col}{row}".format(
                        get_column_letter(table2_offset + len(drivers) + 1),
                        box_count_col="B",
                        row=group_rows_start,
                    )
                ],
                "percent": [
                    "={}{row}/{shipment_cost_col}{row}".format(
                        get_column_letter(table2_offset + len(drivers) + 1),
                        shipment_cost_col="C",
                        row=group_rows_start,
                    )
                ],
            }
        )

        table3_header_df = pd.DataFrame(
            {
                **{
                    manager: [
                        "=SUM({manager_col}{}:{manager_col}{})".format(
                            *table1_range,
                            manager_col=get_column_letter(i),
                        )
                    ]
                    for i, manager in enumerate(managers, start=table3_offset + 1)
                },
                "total_payment": [
                    "=SUM({}{row}:{}{row})".format(
                        get_column_letter(table3_offset + 1),
                        get_column_letter(table3_offset + len(managers)),
                        row=group_rows_start,
                    )
                ],
                "box_cost": [
                    "={}{row}/{box_count_col}{row}".format(
                        get_column_letter(table3_offset + len(managers) + 1),
                        box_count_col="B",
                        row=group_rows_start,
                    )
                ],
                "percent": [
                    "={}{row}/{shipment_cost_col}{row}".format(
                        get_column_letter(table3_offset + len(managers) + 1),
                        shipment_cost_col="C",
                        row=group_rows_start,
                    )
                ],
            }
        )

        for r1, r2, r3 in zip(
            dataframe_to_rows(table1_header_df, index=False, header=False),
            dataframe_to_rows(table2_header_df, index=False, header=False),
            dataframe_to_rows(table3_header_df, index=False, header=False),
        ):
            ws.append(r1 + [None] + r2 + [None] + r3)
            for cell in list(ws)[-1]:
                cell.style = "subheaders_style"

        for r1, r2, r3 in zip(
            dataframe_to_rows(table1_df, index=False, header=False),
            dataframe_to_rows(table2_df, index=False, header=False),
            dataframe_to_rows(table3_df, index=False, header=False),
        ):
            ws.append(r1 + [None] + r2 + [None] + r3)
            for cell in list(ws)[-1]:
                cell.style = "general_style"

        ws.row_dimensions.group(
            group_rows_start + 1,
            group_rows_start + table1_size,
            hidden=True,
            outline_level=1,
        )

        # move to next sub-header row
        group_rows_start += len(table1_header_df) + len(table1_df)

    # setup footer cells for tables
    ws.cell(row=group_rows_start, column=1).value = "ИТОГО:"
    ws.cell(row=group_rows_start, column=2).value = (
        "=SUM({box_count_col}{}:{box_count_col}{}) / 2".format(
            2, group_rows_start - 1, box_count_col="B"
        )
    )
    ws.cell(row=group_rows_start, column=3).value = (
        "=SUM({box_count_col}{}:{box_count_col}{}) / 2".format(
            2, group_rows_start - 1, box_count_col="C"
        )
    )
    for i in range(1, 3 + 1):
        ws.cell(row=group_rows_start, column=i).style = "footer_style"

    table2_offset = 4
    for i, driver in enumerate(
        list(drivers) + ["ВОДИТЕЛИ", "СТОИМОСТЬ КОРОБКИ", "ПРОЦЕНТ"], start=1
    ):
        ws.cell(row=group_rows_start, column=table2_offset + i).value = (
            "=SUM({driver_col}{}:{driver_col}{}) / 2".format(
                2,
                group_rows_start - 1,
                driver_col=get_column_letter(table2_offset + i),
            )
        )
        ws.cell(row=group_rows_start, column=table2_offset + i).style = "footer_style"
    ws.cell(row=group_rows_start, column=table2_offset + len(drivers) + 2).value = (
        "={total_payment_col}{row}/{total_box_count_col}{row}".format(
            row=group_rows_start,
            total_box_count_col="B",
            total_payment_col=get_column_letter(table2_offset + len(drivers) + 1),
        )
    )
    ws.cell(row=group_rows_start, column=table2_offset + len(drivers) + 3).value = (
        "={total_payment_col}{row}/{total_shipment_cost_col}{row}".format(
            row=group_rows_start,
            total_shipment_cost_col="C",
            total_payment_col=get_column_letter(table2_offset + len(drivers) + 1),
        )
    )

    table3_offset = table2_offset + len(drivers) + 3 + 1
    for i, manager in enumerate(
        list(managers) + ["ДЕВОЧКИ", "СТОИМОСТЬ КОРОБКИ", "ПРОЦЕНТ"], start=1
    ):
        ws.cell(row=group_rows_start, column=table3_offset + i).value = (
            "=SUM({manager_col}{}:{manager_col}{}) / 2".format(
                2,
                group_rows_start - 1,
                manager_col=get_column_letter(table3_offset + i),
            )
        )
        ws.cell(row=group_rows_start, column=table3_offset + i).style = "footer_style"

        ws.cell(row=group_rows_start, column=table3_offset + i).value = (
            "=SUM({manager_col}{}:{manager_col}{}) / 2".format(
                2,
                group_rows_start - 1,
                manager_col=get_column_letter(table3_offset + i),
            )
        )
    ws.cell(row=group_rows_start, column=table3_offset + len(managers) + 2).value = (
        "={total_payment_col}{row}/{total_box_count_col}{row}".format(
            row=group_rows_start,
            total_box_count_col="B",
            total_payment_col=get_column_letter(table3_offset + len(managers) + 1),
        )
    )
    ws.cell(row=group_rows_start, column=table3_offset + len(managers) + 3).value = (
        "={total_payment_col}{row}/{total_shipment_cost_col}{row}".format(
            row=group_rows_start,
            total_shipment_cost_col="C",
            total_payment_col=get_column_letter(table3_offset + len(managers) + 1),
        )
    )

    # last styling before save
    for cell in ws[get_column_letter(table2_offset)]:
        cell.fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
    for cell in ws[get_column_letter(table3_offset)]:
        cell.fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

    for cell in ws[get_column_letter(table2_offset + len(drivers) + 3)]:
        cell.number_format = "0.00%"
    for cell in ws[get_column_letter(table3_offset + len(managers) + 3)]:
        cell.number_format = "0.00%"

    ws.freeze_panes = "D2"

    # Change column width
    cols_width = (
        [60, 25, 25]  # table 1
        + [10]  # Seperator
        + [15] * (len(drivers) + 3)  # table 2
        + [10]  # Seperator
        + [15] * (len(managers) + 3)  # table 3
    )
    dim_holder = DimensionHolder(worksheet=ws)
    for i, col in enumerate(range(ws.min_column, ws.max_column + 1)):
        try:
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=cols_width[i]
            )
        except Exception as _:
            continue
    ws.column_dimensions = dim_holder

    # Change row height
    for i, _ in enumerate(ws.iter_rows(min_row=1, max_row=group_rows_start)):
        ws.row_dimensions[i].height = 25

    buffer = BytesIO()
    workbook.save(buffer)

    return buffer
