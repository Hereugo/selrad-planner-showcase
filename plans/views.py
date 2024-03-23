import io
import re
import json
from urllib.parse import urlencode, quote_plus
from itertools import groupby

import openpyxl
from openpyxl.styles import Border, Side

from django.views.generic import ListView
from django.shortcuts import render, redirect
from django.http import FileResponse

from .filters import PlanFilter
from .forms import PlanForm
from .models import Plan, Worklist
from .serializers import PlanSerializer

from clients.models import Client, Address
from managers.models import Manager


def index(request):
    return render(request, 'index.html')

def aside_buttons(request):
    # get reuqests query parameters
    payload = json.loads(request.GET.get('filter_params', '{}'))

    # convert filter to urlencoded string 
    res = urlencode(payload, quote_via=quote_plus)

    return render(request, 'aside_buttons.html', {'urlencode': res})


class PlanListView(ListView):
    model = Plan
    template_name = 'plans.html'
    queryset = Plan.objects.all()
    filterset_class = PlanFilter
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # remove 'page' from GET parameters, so we can keep the search parameters.
        context['urlencode'] = re.split(
            r'page=\d+', self.request.GET.urlencode())[-1]
        if 'page' in self.request.GET.urlencode():
            context['urlencode'] = context['urlencode'][1:]

        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = self.filterset_class(self.request.GET, queryset=queryset).qs
        return queryset


def plan_create(request, pk=None):
    if request.method == 'POST':
        form = PlanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('plans')

    plan = Plan.objects.filter(pk=pk).first()
    clients = Client.objects.all()
    worklist = Worklist.objects.all()
    managers = Manager.objects.all()

    return render(request, 'plan_modal_form.html', {
        'plan': plan,
        'clients': clients,
        'worklist': worklist,
        'managers': managers
    })


def get_client_address(request):
    client_pk = int(request.GET.get('selected_client') or '-1')
    plan_pk = request.GET.get('plan_pk')

    client = Client.objects.filter(pk=client_pk).first()
    if client:
        addresses = Address.objects.filter(clients__pk__in=[client_pk])
    else:
        addresses = Address.objects.none()

    if plan_pk:
        plan = Plan.objects.filter(pk=plan_pk).first()
    else:
        plan = Plan.objects.none()

    return render(request, 'client_address_form.html', {'addresses': addresses, 'plan': plan, 'client': client})


def plan_show_map(request, pk=None):
    # TODO: Figure out how to send json of m2m fields with all necessary fields
    queryset_plans = Plan.objects.all()
    queryset_plans = PlanFilter(request.GET, queryset=queryset_plans).qs

    data = PlanSerializer(queryset_plans, many=True).data

    return render(request, 'map.html', {'plans_json': data}) 


def plans_excel(request):
    COL_DICT = {
        'assigned_date': 1,
        'client': 2,
        'address': 3,
        'manager': 4,
        'worklist': 5,
        'comment': 6,
        'shipment_cost': 7,
        'box_count': 8
    }

    queryset = Plan.objects.all()
    queryset = PlanFilter(request.GET, queryset=queryset).qs

    buffer = io.BytesIO()

    workbook = openpyxl.load_workbook('./static/docs/standard_plan.xlsx')
    ws = workbook.active

    # Fillin the excel file with the plans

    if 'assigned_date_cell' not in workbook.style_names:
        date_style = openpyxl.styles.NamedStyle(name='assigned_date_cell')
        # set format to date with weekday
        date_style.number_format = 'DD.MM.YYYY'

        # font color white, background color gray and bold
        date_style.font = openpyxl.styles.Font(color='FFFFFF', bold=True)
        date_style.fill = openpyxl.styles.PatternFill(
            start_color='808080', end_color='808080', fill_type='solid')
        date_style.alignment = openpyxl.styles.Alignment(
            horizontal='center', vertical='center')
        date_style.alignment.wrap_text = True

        workbook.add_named_style(date_style)

    if 'general_style' not in workbook.style_names:
        general_style = openpyxl.styles.NamedStyle(name='general_style')
        general_style.alignment.wrap_text = True

        general_style.border = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))

        workbook.add_named_style(general_style)

    # get earliest and latest dates from the queryset
    earliest_date = queryset.earliest('assigned_date').assigned_date
    latest_date = queryset.latest('assigned_date').assigned_date

    ws.cell(row=1, column=1).value = f'Планы на {earliest_date} - {latest_date}'
    ws.cell(row=1, column=1).style = 'assigned_date_cell'

    row = 2
    for assigned_date, plans in groupby(queryset, key=lambda p: p.assigned_date):
        for i, plan in enumerate(plans, start=1):
            ws.cell(row=row + i, column=COL_DICT['client']).value = plan.client.name
            ws.cell(row=row + i, column=COL_DICT['address']).value = plan.address.street
            ws.cell(row=row + i, column=COL_DICT['manager']).value = ', '.join(
                [str(m) for m in plan.managers.all()])
            ws.cell(row=row + i, column=COL_DICT['worklist']).value = ', '.join(
                [str(w) for w in plan.worklist.all()])
            ws.cell(row=row + i,
                    column=COL_DICT['comment']).value = plan.comment
            ws.cell(
                row=row + i, column=COL_DICT['shipment_cost']).value = plan.shipment_cost
            ws.cell(row=row + i,
                    column=COL_DICT['box_count']).value = plan.box_count

            # apply the general style to the row
            for col in range(2, 9):
                ws.cell(row=row + i, column=col).style = 'general_style'

        x = queryset.filter(assigned_date=assigned_date).count() + 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row + x, end_column=1)
        ws.cell(row=row, column=1).value = assigned_date
        ws.cell(row=row, column=1).style = 'assigned_date_cell'

        row += x + 1

    workbook.save(buffer)
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename='plans.xlsx')
